"""
Sistema de Venta y Validacion de Entradas con Codigos QR — Multi-Evento
=======================================================================
Backend: Flask + Supabase (PostgreSQL serverless)
Deploy: Vercel (funcion serverless)

Modelo de usuarios:
  - La tabla users es GLOBAL: rol 'admin' (administra todo) o NULL (staff).
  - El staff se asigna a eventos via evento_usuarios (evento_id, usuario_id, rol).
    Un usuario puede estar en varios eventos con roles distintos.
  - Al iniciar sesion:
      * admin -> va directo al panel de administracion (dashboard global).
      * staff con 1 evento activo -> entra directo a vendedor/portero.
      * staff con varios eventos -> elige evento en /elegir.
  - El admin NO opera: no vende ni valida, solo administra.
  - Todos los datos (entradas, logs) se filtran por el evento activo en sesion.
  - Cedula validada en formato costa-ricense: X-XXXX-XXXX (ej. 6-0510-0347)
"""

import os
import re
import time
import uuid
import io
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from functools import wraps
from flask import (Flask, request, jsonify, render_template,
                   session, redirect, url_for, Response)
from werkzeug.security import check_password_hash, generate_password_hash
from supabase import create_client, Client

# Cargar .env en desarrollo local (Vercel usa env vars configuradas en el dashboard)
try:
    from load_env import cargar_env
    cargar_env()
except ImportError:
    pass

# -----------------------------------------------------------
# Configuracion Supabase
# -----------------------------------------------------------
SUPABASE_URL = os.environ.get("SUPABASE_URL", "https://bibdstpwmtfsvbcduvey.supabase.co")

# Service role key: usado server-side (bypasses RLS). Se lee de env
# para no hardcodearlo en el repositorio. Sin fallback: con RLS activa,
# usar otra key fallaria en silencio; mejor arrancar con error claro.
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_KEY", "")
if not SUPABASE_KEY:
    raise RuntimeError(
        "Falta SUPABASE_SERVICE_KEY en las variables de entorno. "
        "Configurala en Vercel (production) o en .env local.")

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# -----------------------------------------------------------
# Configuracion Flask
# -----------------------------------------------------------
app = Flask(__name__,
            template_folder=os.path.join(os.path.dirname(os.path.abspath(__file__)), "templates"))

app.secret_key = os.environ.get("SECRET_KEY", "clave-local-de-desarrollo")

# Cookies de sesion: solo por HTTPS y sin CSRF cross-site
app.config["SESSION_COOKIE_SECURE"] = True
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"

# Nombre por defecto de la pagina (no atado a ningun evento en particular)
NOMBRE_EVENTO_DEFAULT = os.environ.get("NOMBRE_EVENTO", "Sistema de Entradas")

# Tablas
TABLA = "entradas"
TABLA_USERS = "users"
TABLA_LOGS = "logs"
TABLA_EVENTOS = "eventos"
TABLA_ASIGNACIONES = "evento_usuarios"
TABLA_COLEGIOS = "colegios"

# Regex: cedula costa-ricense X-XXXX-XXXX (ej. 6-0510-0347)
CEDULA_CR = re.compile(r'^\d{1,2}-\d{4}-\d{4}$')


# -----------------------------------------------------------
# Context processor — variables globales en templates
# -----------------------------------------------------------
@app.context_processor
def inject_evento():
    es_admin = session.get("rol") == "admin"
    return {
        "nombre_evento": ("Panel de administración" if es_admin
                          else session.get("evento_nombre", NOMBRE_EVENTO_DEFAULT)),
        "evento_id": session.get("evento_id"),
        "precio_entrada": session.get("precio_entrada", 0),
        "rol": session.get("rol", ""),
        "usuario": session.get("usuario", ""),
        "multi_evento": session.get("multi_evento", False),
    }


# -----------------------------------------------------------
# Helpers
# -----------------------------------------------------------
def validar_cedula(cedula):
    """Valida formato de cedula costa-ricense: X-XXXX-XXXX"""
    return bool(CEDULA_CR.match(cedula.strip()))


def evento_acts():
    """Devuelve el evento activo de la session, o None si no hay uno seleccionado."""
    eid = session.get("evento_id")
    if eid:
        return {
            "id": eid,
            "nombre": session.get("evento_nombre", ""),
            "precio": session.get("precio_entrada", 0),
        }
    return None


def registrar_log(accion, detalle=""):
    """Inserta una accion en la tabla de logs."""
    try:
        supabase.table(TABLA_LOGS).insert({
            "accion": accion,
            "detalle": detalle[:500],
            "usuario": session.get("usuario", "?"),
            "evento_id": session.get("evento_id"),
        }).execute()
    except Exception:
        pass


def generar_codigo():
    """Genera un codigo alfanumerico unico de 8 caracteres."""
    return uuid.uuid4().hex[:8].upper()


def entrada_duplicada(nombre, cedula, evento_id):
    """
    Devuelve 'nombre', 'cedula' o None si ya existe una entrada con esos
    datos DENTRO DEL MISMO EVENTO.
    """
    def escapar(v):
        return v.replace('%', r'\%').replace('_', r'\_')

    r1 = supabase.table(TABLA).select("id").eq("evento_id", evento_id) \
        .ilike("nombre", escapar(nombre)).limit(1).execute()
    if r1.data:
        return "nombre"
    r2 = supabase.table(TABLA).select("id").eq("evento_id", evento_id) \
        .ilike("cedula", escapar(cedula)).limit(1).execute()
    if r2.data:
        return "cedula"
    return None


def asignaciones_usuario(usuario_id):
    """Devuelve las asignaciones activas de un usuario: [{evento_id, nombre, precio, rol}]."""
    resp = supabase.table(TABLA_ASIGNACIONES) \
        .select("evento_id, rol, eventos(nombre, precio_entrada, activo)") \
        .eq("usuario_id", usuario_id).execute()
    eventos = []
    for a in resp.data or []:
        ev = a.get("eventos") or {}
        if ev.get("activo", True):
            eventos.append({
                "evento_id": a["evento_id"],
                "nombre": ev.get("nombre", ""),
                "precio_entrada": ev.get("precio_entrada", 0),
                "rol": a["rol"],
            })
    return sorted(eventos, key=lambda e: e["evento_id"])


# -----------------------------------------------------------
# Control de acceso
# -----------------------------------------------------------
def require_rol(*roles):
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if not session.get("auth"):
                if request.path.startswith("/api"):
                    return jsonify({"ok": False, "error": "No autorizado"}), 401
                return redirect(url_for("login"))
            if session.get("rol") not in roles and \
                    not (session.get("multi_evento") and "elegir" in roles):
                if request.path.startswith("/api"):
                    return jsonify({"ok": False, "error": "Rol sin permiso"}), 401
                destino = session.get("rol")
                if destino in ("vendedor", "portero", "admin"):
                    return redirect(url_for(destino))
                if destino == "elegir":
                    return redirect(url_for("elegir"))
                return redirect(url_for("login"))
            return f(*args, **kwargs)
        return wrapper
    return decorator


def require_evento():
    """True si hay un evento activo en session."""
    return session.get("evento_id") is not None


def es_admin_general():
    """Admin sin colegio asignado: controla todos los colegios."""
    return not session.get("colegio_id")


def eventos_visibles(desc=False):
    """Eventos alcanzables por el admin en sesión (su colegio o todos).

    Si la migración de colegios aún no se ejecutó, devuelve todos.
    """
    def _consultar(filtrar):
        q = supabase.table(TABLA_EVENTOS).select(
            "id,nombre,precio_entrada,activo,colegio_id")
        if filtrar:
            q = q.eq("colegio_id", session["colegio_id"])
        return q.order("id", desc=desc).execute()

    if session.get("colegio_id"):
        try:
            return _consultar(True).data or []
        except Exception:
            pass  # columna colegio_id inexistente todavía
    try:
        return _consultar(False).data or []
    except Exception:
        return []


def usuario_en_alcance(user_row):
    """True si el usuario pertenece al alcance del admin en sesión.

    Alcance del admin de colegio: usuarios de su colegio y staff
    asignado a eventos de su colegio. El admin general alcanza a todos.
    """
    if not session.get("colegio_id"):
        return True
    if user_row.get("colegio_id") == session.get("colegio_id"):
        return True
    ev_ids = {e["id"] for e in eventos_visibles()}
    if not ev_ids:
        return False
    asig = supabase.table(TABLA_ASIGNACIONES).select("evento_id") \
        .eq("usuario_id", user_row["id"]).execute().data or []
    return any(a["evento_id"] in ev_ids for a in asig)


def _usuario_por_id(user_id):
    """Fila de usuario con colegio_id si existe la columna."""
    try:
        r = supabase.table(TABLA_USERS).select(
            "id,usuario,nombre,rol,colegio_id").eq("id", user_id).execute()
    except Exception:
        r = supabase.table(TABLA_USERS).select(
            "id,usuario,nombre,rol").eq("id", user_id).execute()
        for u in r.data or []:
            u.setdefault("colegio_id", None)
    return r.data[0] if r.data else None


# -----------------------------------------------------------
# Rutas de interfaz (protegidas con login)
# -----------------------------------------------------------
@app.route("/login")
def login():
    return render_template("login.html")


@app.route("/")
@require_rol("vendedor", "portero", "admin", "elegir")
def index():
    destino = session.get("rol", "vendedor")
    if destino == "admin":
        return redirect(url_for("admin"))
    if destino == "elegir":
        return redirect(url_for("elegir"))
    return redirect(url_for(destino))


@app.route("/elegir")
@require_rol("elegir")
def elegir():
    return render_template("elegir.html")


@app.route("/vendedor")
@require_rol("vendedor")
def vendedor():
    if not require_evento():
        return redirect(url_for("elegir"))
    return render_template("vendedor.html", rol=session.get("rol"))


@app.route("/portero")
@require_rol("portero")
def portero():
    if not require_evento():
        return redirect(url_for("elegir"))
    return render_template("portero.html", rol=session.get("rol"))


@app.route("/centro")
@require_rol("vendedor", "portero")
def centro():
    if not require_evento():
        return redirect(url_for("elegir"))
    return render_template("centro.html", rol=session.get("rol"))


@app.route("/admin")
@require_rol("admin")
def admin():
    colegio_nombre = ""
    if session.get("colegio_id"):
        try:
            colegio_nombre = supabase.table(TABLA_COLEGIOS).select("nombre") \
                .eq("id", session["colegio_id"]).single().execute().data["nombre"]
        except Exception:
            pass
    return render_template("admin.html", rol=session.get("rol"),
                           es_general=not session.get("colegio_id"),
                           colegio_nombre=colegio_nombre)


# -----------------------------------------------------------
# API: login / logout
# -----------------------------------------------------------
# ponytail: rate limit en memoria (se reinicia por instancia serverless);
# igual frena la fuerza bruta. Redis si algun dia importa multi-instancia.
_LOGIN_INTENTOS: dict = {}

def _login_bloqueado():
    ip = request.headers.get("x-forwarded-for", request.remote_addr or "?").split(",")[0]
    ahora = time.time()
    ventana = [t for t in _LOGIN_INTENTOS.get(ip, []) if ahora - t < 300]
    _LOGIN_INTENTOS[ip] = ventana
    if len(ventana) >= 10:
        return True
    return False

def _registrar_intento():
    ip = request.headers.get("x-forwarded-for", request.remote_addr or "?").split(",")[0]
    _LOGIN_INTENTOS.setdefault(ip, []).append(time.time())

@app.route("/api/login", methods=["POST"])
def api_login():
    if _login_bloqueado():
        return jsonify({"ok": False,
                        "error": "Demasiados intentos. Esperá unos minutos."}), 429

    data = request.get_json(silent=True)
    if not data or not data.get("usuario") or not data.get("password"):
        return jsonify({"ok": False, "error": "Usuario y contraseña son requeridos"}), 400

    usuario = data["usuario"].strip().lower()
    try:
        resp = supabase.table(TABLA_USERS).select(
            "id,usuario,password_hash,rol,nombre,colegio_id"
        ).eq("usuario", usuario).execute()
    except Exception:
        # La columna colegio_id puede no existir aún (migración pendiente)
        try:
            resp = supabase.table(TABLA_USERS).select(
                "id,usuario,password_hash,rol,nombre"
            ).eq("usuario", usuario).execute()
            for u in resp.data:
                u.setdefault("colegio_id", None)
        except Exception:
            return jsonify({"ok": False,
                            "error": "La tabla users no existe — ejecutá schema.sql en Supabase"}), 500

    if not resp.data:
        _registrar_intento()
        return jsonify({"ok": False, "error": "Usuario o contraseña incorrectos"}), 401

    user = resp.data[0]
    if not check_password_hash(user["password_hash"], data["password"]):
        _registrar_intento()
        registrar_log("login_fallido", usuario)
        return jsonify({"ok": False, "error": "Usuario o contraseña incorrectos"}), 401

    session["auth"] = True
    session["usuario"] = user["usuario"]
    session["nombre"] = user.get("nombre") or ""
    session["multi_evento"] = False
    session["colegio_id"] = user.get("colegio_id")

    # Admin global: entra directo, no opera sobre eventos
    if user["rol"] == "admin":
        session["rol"] = "admin"
        session["evento_id"] = None
        session["evento_nombre"] = None
        session["precio_entrada"] = 0
        destino = "admin"
        alcance = "general" if not user.get("colegio_id") else "colegio"
        registrar_log("login", f"{user.get('nombre', '')} ({user['usuario']}, admin {alcance})")
        return jsonify({"ok": True, "rol": "admin", "destino": "/admin",
                        "usuario": user["usuario"], "nombre": user.get("nombre", ""),
                        "es_general": not user.get("colegio_id")})

    # Staff: buscar asignaciones en eventos activos
    asignaciones = asignaciones_usuario(user["id"])
    if not asignaciones:
        return jsonify({"ok": False,
                        "error": "Este usuario no está asignado a ningún evento activo"}), 401

    if len(asignaciones) == 1:
        a = asignaciones[0]
        session["rol"] = a["rol"]
        session["evento_id"] = a["evento_id"]
        session["evento_nombre"] = a["nombre"]
        session["precio_entrada"] = a["precio_entrada"]
        destino = a["rol"]
        registrar_log("login", f"{user.get('nombre', '')} ({user['usuario']}, "
                               f"{a['rol']} · {a['nombre']})")
        return jsonify({"ok": True, "rol": a["rol"], "destino": f"/{a['rol']}",
                        "usuario": user["usuario"], "nombre": user.get("nombre", ""),
                        "evento_id": a["evento_id"], "evento_nombre": a["nombre"]})

    # Varios eventos: elegir
    session["rol"] = "elegir"
    session["evento_id"] = None
    session["evento_nombre"] = None
    session["precio_entrada"] = 0
    session["multi_evento"] = True
    registrar_log("login", f"{user.get('nombre', '')} ({user['usuario']}, multi-evento)")
    return jsonify({"ok": True, "rol": "elegir", "destino": "/elegir",
                    "usuario": user["usuario"], "nombre": user.get("nombre", "")})


@app.route("/api/logout", methods=["POST"])
def api_logout():
    registrar_log("logout", session.get("usuario", "?"))
    session.clear()
    return jsonify({"ok": True})


# -----------------------------------------------------------
# API: eventos (admin CRUD)
# -----------------------------------------------------------
@app.route("/api/eventos")
@require_rol("admin")
def api_eventos_listar():
    return jsonify({"eventos": eventos_visibles(desc=True)})


@app.route("/api/eventos", methods=["POST"])
@require_rol("admin")
def api_eventos_crear():
    # La creación de eventos es del admin de colegio; el general solo supervisa
    if es_admin_general():
        return jsonify({"ok": False,
                        "error": "Los eventos los crean los admins de cada colegio"}), 403

    data = request.get_json(silent=True) or {}
    nombre = (data.get("nombre") or "").strip()[:100]
    try:
        precio = int(data.get("precio", 1000))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "Precio inválido"}), 400

    if not nombre:
        return jsonify({"ok": False, "error": "Nombre del evento es requerido"}), 400
    if precio < 0:
        return jsonify({"ok": False, "error": "El precio no puede ser negativo"}), 400

    try:
        payload = {"nombre": nombre, "precio_entrada": precio}
        # Admin de colegio crea para SU colegio; el general elige (o deja global)
        colegio_id = session.get("colegio_id") or data.get("colegio_id")
        if colegio_id:
            payload["colegio_id"] = int(colegio_id)
        try:
            supabase.table(TABLA_EVENTOS).insert(payload).execute()
        except Exception as e:
            # Solo degradar si la columna realmente no existe (migración pendiente)
            if "colegio_id" in payload and ("PGRST204" in str(e) or "Could not find the 'colegio_id' column" in str(e)):
                payload.pop("colegio_id")
                supabase.table(TABLA_EVENTOS).insert(payload).execute()
            else:
                raise
        registrar_log("evento_crear", f"{nombre} | precio {precio}")
        return jsonify({"ok": True, "nombre": nombre, "precio": precio})
    except Exception:
        return jsonify({"ok": False, "error": "Ya existe un evento con ese nombre"}), 409


@app.route("/api/eventos/<int:evento_id>", methods=["PUT"])
@require_rol("admin")
def api_evento_editar(evento_id):
    data = request.get_json(silent=True) or {}
    cambios = {}

    if "nombre" in data:
        nombre = (data.get("nombre") or "").strip()[:100]
        if not nombre:
            return jsonify({"ok": False, "error": "Nombre no puede estar vacío"}), 400
        cambios["nombre"] = nombre

    if "precio_entrada" in data:
        try:
            precio = int(data["precio_entrada"])
        except (TypeError, ValueError):
            return jsonify({"ok": False, "error": "Precio inválido"}), 400
        if precio < 0:
            return jsonify({"ok": False, "error": "El precio no puede ser negativo"}), 400
        cambios["precio_entrada"] = precio

    if "activo" in data:
        cambios["activo"] = bool(data["activo"])

    if not cambios:
        return jsonify({"ok": False, "error": "Nada que editar"}), 400

    try:
        resp = supabase.table(TABLA_EVENTOS).update(cambios).eq("id", evento_id).execute()
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 409

    # Actualizar session si es el evento activo
    if session.get("evento_id") == evento_id:
        if "precio_entrada" in cambios:
            session["precio_entrada"] = cambios["precio_entrada"]
        if "nombre" in cambios:
            session["evento_nombre"] = cambios["nombre"]

    detalle = ", ".join(f"{k}: {v}" for k, v in cambios.items())
    registrar_log("evento_editar", f"id {evento_id} · {detalle}")
    return jsonify({"ok": True, "precio": cambios.get("precio_entrada")})


@app.route("/api/eventos/<int:evento_id>", methods=["DELETE"])
@require_rol("admin")
def api_evento_borrar(evento_id):
    if session.get("evento_id") == evento_id:
        session.pop("evento_id", None)
        session.pop("evento_nombre", None)
        session.pop("precio_entrada", None)

    try:
        resp = supabase.table(TABLA_EVENTOS).select("nombre").eq("id", evento_id).execute()
        if not resp.data:
            return jsonify({"ok": False, "error": "Evento no encontrado"}), 404
        nombre = resp.data[0]["nombre"]
        supabase.table(TABLA_EVENTOS).delete().eq("id", evento_id).execute()
        registrar_log("evento_borrar", f"{nombre} (id {evento_id})")
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/mis_eventos")
@require_rol("elegir")
def api_mis_eventos():
    """Eventos activos del usuario logueado, para la pantalla /elegir."""
    user = supabase.table(TABLA_USERS).select("id").eq("usuario", session["usuario"]).execute()
    if not user.data:
        return jsonify({"ok": False, "error": "Usuario no encontrado"}), 404
    return jsonify({"eventos": asignaciones_usuario(user.data[0]["id"])})


@app.route("/api/evento/select", methods=["POST"])
@require_rol("elegir")
def api_evento_select():
    """El staff elige el evento al que quiere entrar (solo los asignados)."""
    data = request.get_json(silent=True) or {}
    evento_id = data.get("evento_id")
    if not evento_id:
        return jsonify({"ok": False, "error": "evento_id requerido"}), 400

    user = supabase.table(TABLA_USERS).select("id").eq("usuario", session["usuario"]).execute()
    if not user.data:
        return jsonify({"ok": False, "error": "Usuario no encontrado"}), 404

    asignaciones = asignaciones_usuario(user.data[0]["id"])
    elegido = next((a for a in asignaciones if a["evento_id"] == int(evento_id)), None)
    if not elegido:
        return jsonify({"ok": False, "error": "No estás asignado a ese evento"}), 403

    session["rol"] = elegido["rol"]
    session["evento_id"] = elegido["evento_id"]
    session["evento_nombre"] = elegido["nombre"]
    session["precio_entrada"] = elegido["precio_entrada"]
    registrar_log("evento_seleccionar", f"{elegido['nombre']} ({elegido['rol']})")
    return jsonify({"ok": True, "destino": f"/{elegido['rol']}",
                    "evento": {"id": elegido["evento_id"], "nombre": elegido["nombre"],
                               "precio_entrada": elegido["precio_entrada"]}})


# -----------------------------------------------------------
# API: generar entrada
# -----------------------------------------------------------
@app.route("/api/generar", methods=["POST"])
@require_rol("vendedor")
def api_generar():
    evento_id = session.get("evento_id")
    if not evento_id:
        return jsonify({"ok": False, "error": "Seleccioná un evento"}), 400

    data = request.get_json(silent=True) or {}
    nombre = (data.get("nombre") or "").strip()[:80]
    cedula = (data.get("cedula") or "").strip()[:30]

    if not nombre or not cedula:
        return jsonify({"ok": False, "error": "Nombre y cédula son requeridos"}), 400

    if not validar_cedula(cedula):
        return jsonify({"ok": False,
                        "error": "Formato de cédula inválido. Usá: X-XXXX-XXXX (ej. 6-0510-0347)"}), 400

    dup = entrada_duplicada(nombre, cedula, evento_id)
    if dup:
        campo = "nombre" if dup == "nombre" else "cédula"
        return jsonify({"ok": False, "error": f"Ya existe una entrada con ese {campo}"}), 409

    precio = session.get("precio_entrada", 0)
    for _ in range(5):
        codigo = generar_codigo()
        try:
            resp = supabase.table(TABLA).insert({
                "codigo": codigo,
                "usado": False,
                "nombre": nombre,
                "cedula": cedula,
                "evento_id": evento_id,
                "precio": precio,
                "vendedor": session.get("usuario", ""),
            }).execute()
            if resp.data:
                registrar_log("venta", f"{nombre} | cédula {cedula} | código {codigo}")
                return jsonify({"ok": True, "codigo": codigo, "id": resp.data[0]["id"],
                                "nombre": nombre, "cedula": cedula, "precio": precio}), 201
        except Exception:
            continue

    return jsonify({"ok": False, "error": "No se pudo generar código"}), 500


# -----------------------------------------------------------
# API: validar entrada
# -----------------------------------------------------------
@app.route("/api/validar", methods=["POST"])
@require_rol("portero")
def api_validar():
    evento_id = session.get("evento_id")
    if not evento_id:
        return jsonify({"ok": False, "error": "Seleccioná un evento"}), 400

    data = request.get_json(silent=True)
    if not data or not data.get("code"):
        return jsonify({"ok": False, "error": "Código requerido"}), 400

    codigo = data["code"].strip().upper()
    origen = "cámara" if data.get("origen") == "camara" else "manual"

    try:
        resp = supabase.table(TABLA).select("usado,nombre,cedula") \
            .eq("evento_id", evento_id).eq("codigo", codigo).execute()
    except Exception as e:
        return jsonify({"ok": False, "error": f"Error de base de datos: {e}"}), 500

    if not resp.data:
        registrar_log("escaneo", f"{codigo} — {origen} | código no encontrado")
        return jsonify({"ok": False, "estado": "inexistente",
                        "error": "Código no encontrado"}), 404

    if resp.data[0]["usado"]:
        registrar_log("escaneo", f"{codigo} — {origen} | ya usado ({resp.data[0].get('nombre', '')})")
        return jsonify({"ok": False, "estado": "usado",
                        "error": "Código ya fue utilizado"}), 409

    supabase.table(TABLA).update({"usado": True}) \
        .eq("evento_id", evento_id).eq("codigo", codigo).execute()
    registrar_log("escaneo", f"{codigo} — {origen} | válido, {resp.data[0].get('nombre', '')}")

    return jsonify({"ok": True, "estado": "valido", "codigo": codigo,
                    "nombre": resp.data[0].get("nombre", ""),
                    "cedula": resp.data[0].get("cedula", "")}), 200


# -----------------------------------------------------------
# API: contador de entradas generadas
# -----------------------------------------------------------
@app.route("/api/contador")
@require_rol("vendedor", "portero")
def api_contador():
    evento_id = session.get("evento_id")
    if not evento_id:
        return jsonify({"total": 0})
    try:
        resp = supabase.table(TABLA).select("id", count="exact") \
            .eq("evento_id", evento_id).execute()
        return jsonify({"total": resp.count})
    except Exception:
        return jsonify({"total": 0})


# -----------------------------------------------------------
# API: centro de datos (stats)
# -----------------------------------------------------------
@app.route("/api/stats")
@require_rol("vendedor", "portero")
def api_stats():
    evento_id = session.get("evento_id")
    precio = session.get("precio_entrada", 0)

    if not evento_id:
        return jsonify({"total": 0, "usadas": 0, "pendientes": 0,
                        "precio": precio, "recaudado_total": 0,
                        "recaudado_usadas": 0, "recaudado_pendientes": 0})

    try:
        ev = supabase.table(TABLA_EVENTOS).select("nombre,precio_entrada") \
            .eq("id", evento_id).execute()
        if ev.data:
            precio = ev.data[0]["precio_entrada"]
            session["precio_entrada"] = precio

        total = supabase.table(TABLA).select("id", count="exact") \
            .eq("evento_id", evento_id).execute().count
        usadas = supabase.table(TABLA).select("id", count="exact") \
            .eq("evento_id", evento_id).eq("usado", True).execute().count
        pendientes = total - usadas

        return jsonify({
            "total": total, "usadas": usadas, "pendientes": pendientes,
            "precio": precio,
            "recaudado_total": total * precio,
            "recaudado_usadas": usadas * precio,
            "recaudado_pendientes": pendientes * precio,
        })
    except Exception:
        return jsonify({"total": 0, "usadas": 0, "pendientes": 0,
                        "precio": precio, "recaudado_total": 0,
                        "recaudado_usadas": 0, "recaudado_pendientes": 0})


@app.route("/api/listar")
@require_rol("vendedor", "portero")
def api_listar():
    evento_id = session.get("evento_id")
    if not evento_id:
        return jsonify({"entradas": []})
    try:
        resp = supabase.table(TABLA).select(
            "id,codigo,usado,creado_en,nombre,cedula,precio,vendedor"
        ).eq("evento_id", evento_id).order("id", desc=True).limit(500).execute()
        return jsonify({"entradas": resp.data})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/reset", methods=["POST"])
@require_rol("vendedor", "portero")
def api_reset():
    evento_id = session.get("evento_id")
    if not evento_id:
        return jsonify({"ok": False, "error": "Seleccioná un evento"}), 400

    data = request.get_json(silent=True)
    if not data or not (data.get("code") or data.get("codigo")):
        return jsonify({"ok": False, "error": "Código requerido"}), 400

    codigo = (data.get("code") or data.get("codigo")).strip().upper()
    resp = supabase.table(TABLA).update({"usado": False}) \
        .eq("evento_id", evento_id).eq("codigo", codigo).execute()

    if not resp.data:
        return jsonify({"ok": False, "error": "Código no encontrado"}), 404

    registrar_log("revertir_escaneo", f"código {codigo} marcado como no usado")
    return jsonify({"ok": True, "codigo": codigo})


@app.route("/api/borrar_todas", methods=["POST"])
@require_rol("vendedor", "portero")
def api_borrar_todas():
    evento_id = session.get("evento_id")
    if not evento_id:
        return jsonify({"ok": False, "error": "Seleccioná un evento"}), 400

    try:
        antes = supabase.table(TABLA).select("id", count="exact") \
            .eq("evento_id", evento_id).execute().count
        supabase.rpc("reset_entradas", {"p_evento_id": int(evento_id)}).execute()
        registrar_log("borrado_total", f"{antes} entradas eliminadas del evento")
        return jsonify({"ok": True, "mensaje": "Entradas eliminadas"})
    except Exception as e:
        return jsonify({"ok": False,
                        "error": "La funcion reset_entradas no existe en Supabase"}), 500


@app.route("/api/exportar")
@require_rol("vendedor", "portero")
def api_exportar():
    evento_id = session.get("evento_id")
    if not evento_id:
        return jsonify({"ok": False, "error": "Seleccioná un evento"}), 400

    try:
        resp = supabase.table(TABLA).select(
            "id,codigo,usado,creado_en,nombre,cedula,precio,vendedor"
        ).eq("evento_id", evento_id).order("id").execute()
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    registrar_log("exportar_csv", f"{len(resp.data)} entradas exportadas")

    total = len(resp.data)
    usadas = sum(1 for r in resp.data if r["usado"])
    pendientes = total - usadas
    precio = session.get("precio_entrada", 0)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["RESUMEN"])
    writer.writerow(["evento", session.get("evento_nombre", "")])
    writer.writerow(["precio_entrada", precio])
    writer.writerow(["total_generadas", total])
    writer.writerow(["total_usadas", usadas])
    writer.writerow(["total_pendientes", pendientes])
    writer.writerow(["recaudado_total", total * precio])
    writer.writerow(["recaudado_usadas", usadas * precio])
    writer.writerow(["recaudado_pendientes", pendientes * precio])
    writer.writerow([])
    writer.writerow(["id", "codigo", "nombre", "cedula", "usado", "precio", "vendedor", "creado_en"])
    for row in resp.data:
        writer.writerow([row["id"], row["codigo"], row.get("nombre", ""),
                         row.get("cedula", ""), row["usado"],
                         row.get("precio", precio), row.get("vendedor", ""),
                         row["creado_en"]])

    response = Response(output.getvalue(), mimetype="text/csv")
    response.headers["Content-Disposition"] = "attachment; filename=entradas.csv"
    return response


@app.route("/api/exportar/excel")
def api_exportar_excel():
    """Excel estilizado Taquilla — con colores por colegio, orden y resumen."""
    if not session.get("auth"):
        return jsonify({"ok": False, "error": "No autorizado"}), 401

    rol = session.get("rol")
    # Determinar alcance según rol
    if rol == "admin":
        eventos = eventos_visibles(desc=False)
        ev_ids = [e["id"] for e in eventos]
        # Filtros opcionales: colegio (general) y evento (general/cole)
        filtro_colegio = request.args.get("colegio_id", type=int)
        if filtro_colegio is not None and es_admin_general():
            ev_ids = [e["id"] for e in eventos if e.get("colegio_id") == filtro_colegio]
            eventos = [e for e in eventos if e.get("colegio_id") == filtro_colegio]
        filtro_evento = request.args.get("evento_id", type=int)
        if filtro_evento is not None:
            if filtro_evento not in ev_ids:
                return jsonify({"ok": False, "error": "Evento no encontrado o sin permiso"}), 404
            ev_ids = [filtro_evento]
            eventos = [e for e in eventos if e["id"] == filtro_evento]
        if not ev_ids:
            ev_ids = [-1]  # para que no traiga nada si no hay eventos
    else:
        # vendedor / portero: solo su evento activo
        evento_id = session.get("evento_id")
        if not evento_id:
            return jsonify({"ok": False, "error": "Seleccioná un evento"}), 400
        ev_ids = [evento_id]
        eventos = eventos_visibles(desc=False)
        filtro_colegio = None

    # Datos de colegios para colores
    try:
        cols = supabase.table(TABLA_COLEGIOS).select("id,nombre,color").execute().data or []
    except Exception:
        cols = []
    mapa_colegio = {c["id"]: c for c in cols}
    mapa_evento = {e["id"]: e for e in eventos}
    # Fallback para eventos que no están en eventos_visibles pero sí en entradas (por si acaso)
    try:
        ev_rows = supabase.table(TABLA_EVENTOS).select("id,nombre,colegio_id,precio_entrada").in_("id", ev_ids).execute().data or []
        for er in ev_rows:
            if er["id"] not in mapa_evento:
                mapa_evento[er["id"]] = er
    except Exception:
        pass

    # Traer entradas
    try:
        # Supabase limita a 1000 por defecto, paginar si hace falta
        entradas = []
        offset = 0
        while True:
            batch = supabase.table(TABLA).select(
                "id,codigo,nombre,cedula,usado,precio,vendedor,creado_en,evento_id"
            ).in_("evento_id", ev_ids).order("id").range(offset, offset+999).execute().data or []
            entradas.extend(batch)
            if len(batch) < 1000:
                break
            offset += 1000
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    # Ordenar por colegio > evento > usado > nombre (para que se vea ordenado)
    def sort_key(r):
        ev = mapa_evento.get(r["evento_id"], {})
        col_id = ev.get("colegio_id")
        col_nombre = mapa_colegio.get(col_id, {}).get("nombre", "Global") if col_id else "Global"
        return (col_nombre, ev.get("nombre",""), 0 if r.get("usado") else 1, r.get("nombre") or "")
    entradas.sort(key=sort_key)

    # Estadísticas
    total = len(entradas)
    usadas = sum(1 for r in entradas if r.get("usado"))
    pendientes = total - usadas
    # Recaudado: sumar precio por entrada (usando precio de la entrada o del evento)
    def precio_de(r):
        return r.get("precio") or mapa_evento.get(r["evento_id"], {}).get("precio_entrada") or 0
    recaudado = sum(precio_de(r) for r in entradas)

    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Entradas"
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # Estilos
    NEGRO = "0B0B0F"
    NEGRO2 = "12121A"
    LIMA = "D9FF3D"
    TEXTO = "ECEBE6"
    SUAVE = "8B8A94"
    BLANCO = "FFFFFF"
    thin_border = Border(
        left=Side(style="thin", color="2A2A30"),
        right=Side(style="thin", color="2A2A30"),
        top=Side(style="thin", color="2A2A30"),
        bottom=Side(style="thin", color="2A2A30"),
    )
    header_fill = PatternFill(start_color=NEGRO, end_color=NEGRO, fill_type="solid")
    header_font = Font(name="Karla", size=8, bold=True, color=LIMA)
    title_font = Font(name="Archivo Black", size=14, bold=True, color=NEGRO)
    title_fill = PatternFill(start_color=LIMA, end_color=LIMA, fill_type="solid")
    sub_font = Font(name="Space Mono", size=8, color=SUAVE)
    cell_font = Font(name="Karla", size=8, color=TEXTO)
    mono_font = Font(name="Space Mono", size=7, color=TEXTO)
    alt_fill = PatternFill(start_color=NEGRO2, end_color=NEGRO2, fill_type="solid")
    even_fill = PatternFill(start_color=NEGRO, end_color=NEGRO, fill_type="solid")

    # Título
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = "TAQUILLA — Reporte de Entradas"
    c.font = title_font
    c.fill = title_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Subtítulo con fecha y alcance
    ws.merge_cells("A2:I2")
    c = ws["A2"]
    alcance_txt = "Todos los colegios" if (rol=="admin" and not session.get("colegio_id") and not filtro_colegio) else (mapa_colegio.get(filtro_colegio, {}).get("nombre") if filtro_colegio else (mapa_colegio.get(session.get("colegio_id"), {}).get("nombre") or "Mi colegio" if rol=="admin" and session.get("colegio_id") else session.get("evento_nombre") or "Reporte"))
    c.value = f"{alcance_txt}  •  {time.strftime('%d/%m/%Y %H:%M')}  •  {total} entradas  •  {usadas} usadas  •  {pendientes} pendientes  •  ₡{recaudado:,}".replace(",", ".")
    c.font = sub_font
    c.fill = PatternFill(start_color="1A1A22", end_color="1A1A22", fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # Resumen en 3 cajitas (fila 3)
    resumen = [
        ("Total vendidas", str(total), "entradas"),
        ("Asistieron", f"{usadas} ({round(usadas/total*100) if total else 0}%)", "usadas"),
        ("Recaudado", f"₡{recaudado:,}".replace(",", "."), "total"),
    ]
    for idx, (label, val, det) in enumerate(resumen, start=1):
        col = (idx-1)*3 + 1
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+2)
        c = ws.cell(row=3, column=col)
        c.value = f"{label}: {val} {det}"
        c.font = Font(name="Space Mono", size=7, bold=True, color=NEGRO)
        # Color por resumen: lima, celeste, gris
        fills = [LIMA, "5AA9FF", "8A8F98"]
        c.fill = PatternFill(start_color=fills[idx-1], end_color=fills[idx-1], fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border
    ws.row_dimensions[3].height = 16

    # Cabecera de tabla (fila 5)
    headers = ["#", "Código", "Nombre", "Cédula", "Evento", "Colegio", "Precio", "Estado", "Vendedor / Fecha"]
    col_widths = [5, 12, 22, 14, 20, 18, 9, 10, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    header_row = 5
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border
    ws.row_dimensions[header_row].height = 20

    # Filas
    row_idx = header_row + 1
    for idx, r in enumerate(entradas, 1):
        ev = mapa_evento.get(r["evento_id"], {})
        col_id = ev.get("colegio_id")
        col_info = mapa_colegio.get(col_id) if col_id else None
        colegio_nombre = col_info["nombre"] if col_info else ("Global" if ev else "—")
        colegio_color = (col_info["color"] if col_info and col_info.get("color") else ("#8A8F98" if colegio_nombre!="Global" else "#D9FF3D")).lstrip("#")
        estado = "Usada" if r.get("usado") else "Pendiente"
        precio = precio_de(r)
        # Color sutil por colegio en la primera columna
        is_alt = idx % 2 == 0
        fill = alt_fill if is_alt else even_fill

        vals = [
            idx,
            r.get("codigo",""),
            r.get("nombre","") or "—",
            r.get("cedula","") or "—",
            ev.get("nombre","") or "—",
            colegio_nombre,
            f"₡{precio:,}".replace(",", "."),
            estado,
            f"{r.get('vendedor','—')}  {r.get('creado_en','')[:10]}",
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.font = mono_font if col in (2,4,7,8) else cell_font
            c.fill = fill
            c.border = thin_border
            c.alignment = Alignment(horizontal="center" if col in (1,7,8) else "left", vertical="center")
            if col == 1:
                # Indicador de colegio
                c.fill = PatternFill(start_color=colegio_color, end_color=colegio_color, fill_type="solid")
                c.font = Font(name="Karla", size=7, bold=True, color=NEGRO if colegio_color.upper() in ("D9FF3D","5AA9FF","FFB84D") else BLANCO)
                c.alignment = Alignment(horizontal="center", vertical="center")
        # Color de estado
        estado_cell = ws.cell(row=row_idx, column=8)
        if r.get("usado"):
            estado_cell.fill = PatternFill(start_color="1E3A28", end_color="1E3A28", fill_type="solid")
            estado_cell.font = Font(name="Karla", size=7, bold=True, color="7EE787")
        else:
            estado_cell.fill = PatternFill(start_color="2A2A1E", end_color="2A2A1E", fill_type="solid")
            estado_cell.font = Font(name="Karla", size=7, bold=True, color="FFD966")
        ws.row_dimensions[row_idx].height = 16
        row_idx += 1

    if total == 0:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=9)
        c = ws.cell(row=row_idx, column=1, value="Sin entradas para este filtro")
        c.font = Font(name="Space Mono", size=9, italic=True, color=SUAVE)
        c.alignment = Alignment(horizontal="center")
        row_idx += 1

    # Totales al pie
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
    c = ws.cell(row=row_idx, column=1, value="TOTALES")
    c.font = Font(name="Karla", size=8, bold=True, color=LIMA)
    c.fill = header_fill
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = thin_border
    for col in range(2,7):
        cc = ws.cell(row=row_idx, column=col)
        cc.fill = header_fill
        cc.border = thin_border
    ws.cell(row=row_idx, column=7, value=f"₡{recaudado:,}".replace(",", ".")).font = Font(name="Space Mono", size=8, bold=True, color=BLANCO)
    ws.cell(row=row_idx, column=7).fill = header_fill
    ws.cell(row=row_idx, column=7).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row_idx, column=7).border = thin_border
    ws.cell(row=row_idx, column=8, value=f"{usadas}/{total}").font = Font(name="Karla", size=8, bold=True, color=BLANCO)
    ws.cell(row=row_idx, column=8).fill = header_fill
    ws.cell(row=row_idx, column=8).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row_idx, column=8).border = thin_border
    ws.cell(row=row_idx, column=9, value=time.strftime("%d/%m/%Y")).font = Font(name="Space Mono", size=7, color=SUAVE)
    ws.cell(row=row_idx, column=9).fill = header_fill
    ws.cell(row=row_idx, column=9).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row_idx, column=9).border = thin_border
    ws.row_dimensions[row_idx].height = 18

    # Ajustes finales
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    registrar_log("exportar_excel", f"{total} filas → {alcance_txt}")

    filename = f"taquilla_{alcance_txt.lower().replace(' ', '_')}_{time.strftime('%Y%m%d')}.xlsx"
    # Limpiar nombre
    filename = re.sub(r"[^a-z0-9_\-\.]", "_", filename)

    return Response(
        out.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# -----------------------------------------------------------
# API: admin — dashboard, eventos, logs, usuarios y asignaciones
# -----------------------------------------------------------
@app.route("/api/dashboard")
@require_rol("admin")
def api_dashboard():
    """Estadisticas generales, por evento y por colegio (admin general).

    Una sola query de entradas (in_ sobre eventos visibles) en vez de
    2-3 queries por evento.
    """
    eventos = eventos_visibles()
    ev_ids = [e["id"] for e in eventos]

    general = {"eventos": len(eventos), "total": 0, "usadas": 0,
               "recaudado_total": 0, "recaudado_usadas": 0}
    por_evento = []
    es_general = es_admin_general()

    # Agregados por evento y vendedor: una sola lectura de entradas
    agg = {e["id"]: {"total": 0, "usadas": 0, "por_vendedor": {}} for e in eventos}
    if ev_ids:
        try:
            rows = supabase.table(TABLA).select("evento_id,usado,vendedor") \
                .in_("evento_id", ev_ids).execute().data or []
            for r in rows:
                a = agg.get(r["evento_id"])
                if not a:
                    continue
                a["total"] += 1
                if r.get("usado"):
                    a["usadas"] += 1
                v = r.get("vendedor") or "—"
                pv = a["por_vendedor"].setdefault(v, {"total": 0, "usadas": 0})
                pv["total"] += 1
                if r.get("usado"):
                    pv["usadas"] += 1
        except Exception as e:
            print(f"[dashboard] error leyendo entradas: {e}")

    for ev in eventos:
        eid = ev["id"]
        precio = ev["precio_entrada"]
        total = agg[eid]["total"]
        usadas = agg[eid]["usadas"]

        por_evento.append({
            "id": eid, "nombre": ev["nombre"], "precio_entrada": precio,
            "activo": ev["activo"],
            "colegio_id": ev.get("colegio_id"),
            "total": total, "usadas": usadas, "pendientes": total - usadas,
            "recaudado_total": total * precio,
            "recaudado_usadas": usadas * precio,
            "por_vendedor": agg[eid]["por_vendedor"],
        })
        general["total"] += total
        general["usadas"] += usadas
        general["recaudado_total"] += total * precio
        general["recaudado_usadas"] += usadas * precio

    respuesta = {"general": general, "por_evento": por_evento}

    # Desglose por colegio (solo admin general)
    if es_general:
        nombres, colores = {}, {}
        try:
            try:
                cols = supabase.table(TABLA_COLEGIOS).select(
                    "id,nombre,color").execute().data or []
            except Exception:
                cols = supabase.table(TABLA_COLEGIOS).select(
                    "id,nombre").execute().data or []
            nombres = {c["id"]: c["nombre"] for c in cols}
            colores = {c["id"]: c.get("color") for c in cols}
        except Exception:
            pass

        por_colegio = {}
        for ev, datos in zip(eventos, por_evento):
            cid = ev.get("colegio_id")
            clave = cid if cid else None
            g = por_colegio.setdefault(clave, {
                "colegio_id": cid,
                "nombre": nombres.get(cid) if cid else "Global",
                "color": colores.get(cid),
                "eventos": 0, "total": 0, "usadas": 0,
                "recaudado_total": 0, "recaudado_usadas": 0})
            g["eventos"] += 1
            g["total"] += datos["total"]
            g["usadas"] += datos["usadas"]
            g["recaudado_total"] += datos["recaudado_total"]
            g["recaudado_usadas"] += datos["recaudado_usadas"]

        lista = sorted(por_colegio.values(),
                       key=lambda x: -x["recaudado_total"])
        # Color de colegio también en el detalle por evento (para gráficas)
        for ev in respuesta["por_evento"]:
            cid = ev.get("colegio_id")
            ev["colegio_color"] = colores.get(cid) if cid else None
        respuesta["por_colegio"] = lista

    return jsonify(respuesta)


@app.route("/api/logs")
@require_rol("admin")
def api_logs():
    def _consultar(filtrar):
        q = supabase.table(TABLA_LOGS).select(
            "id,accion,detalle,usuario,creado_en,evento_id")
        if filtrar:
            ev_ids = [e["id"] for e in eventos_visibles()]
            if not ev_ids:
                return None
            q = q.in_("evento_id", ev_ids)
        return q.order("id", desc=True).limit(300).execute()

    try:
        resp = _consultar(bool(session.get("colegio_id")))
        if resp is None:
            return jsonify({"logs": []})
        return jsonify({"ok": True, "logs": resp.data})
    except Exception:
        # Sin filtro si la migración de colegios no corrió todavía
        try:
            resp = supabase.table(TABLA_LOGS).select(
                "id,accion,detalle,usuario,creado_en,evento_id"
            ).order("id", desc=True).limit(300).execute()
            return jsonify({"ok": True, "logs": resp.data})
        except Exception as e:
            return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/users")
@require_rol("admin")
def api_users_listar():
    try:
        try:
            users = supabase.table(TABLA_USERS).select(
                "id,usuario,rol,nombre,colegio_id,creado_en").order("id").execute().data
        except Exception:
            users = supabase.table(TABLA_USERS).select(
                "id,usuario,rol,nombre,creado_en").order("id").execute().data
            for u in users:
                u.setdefault("colegio_id", None)
        asig = supabase.table(TABLA_ASIGNACIONES).select(
            "usuario_id,evento_id,rol,eventos(nombre)").execute().data
        eventos_nombre = {a["evento_id"]: (a.get("eventos") or {}).get("nombre", "")
                          for a in asig}
        por_usuario = {}
        for a in asig:
            por_usuario.setdefault(a["usuario_id"], []).append({
                "evento_id": a["evento_id"],
                "evento_nombre": eventos_nombre.get(a["evento_id"], ""),
                "rol": a["rol"],
            })
        for u in users:
            u["asignaciones"] = por_usuario.get(u["id"], [])

        # Admin de colegio: solo ve usuarios de su colegio y su staff
        if session.get("colegio_id"):
            ev_ids = {e["id"] for e in eventos_visibles()}
            en_alcance = {u["id"] for u in users
                          if u.get("colegio_id") == session["colegio_id"]}
            for a in asig:
                if a["evento_id"] in ev_ids:
                    en_alcance.add(a["usuario_id"])
            users = [u for u in users if u["id"] in en_alcance]

        return jsonify({"users": users})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/users", methods=["POST"])
@require_rol("admin")
def api_users_crear():
    data = request.get_json(silent=True) or {}
    usuario = (data.get("usuario") or "").strip().lower()
    password = data.get("password") or ""
    admin = bool(data.get("admin"))
    nombre = (data.get("nombre") or "").strip()[:60]

    if not usuario or not password:
        return jsonify({"ok": False, "error": "usuario y password son requeridos"}), 400

    existe = supabase.table(TABLA_USERS).select("id") \
        .eq("usuario", usuario).execute()
    if existe.data:
        return jsonify({"ok": False, "error": "Ese usuario ya existe"}), 409

    try:
        payload = {
            "usuario": usuario,
            "password_hash": generate_password_hash(password),
            "rol": "admin" if admin else None,
            "nombre": nombre,
        }
        # Admin de colegio crea dentro de su colegio; el general elige destino
        colegio_id = session.get("colegio_id") or data.get("colegio_id")
        if colegio_id:
            payload["colegio_id"] = int(colegio_id)
        try:
            supabase.table(TABLA_USERS).insert(payload).execute()
        except Exception as e:
            # Solo degradar si la columna realmente no existe (migración pendiente);
            # NUNCA ante errores de FK u otros, que crearían admins generales por accidente
            if "colegio_id" in payload and ("PGRST204" in str(e) or "Could not find the 'colegio_id' column" in str(e)):
                payload.pop("colegio_id")
                supabase.table(TABLA_USERS).insert(payload).execute()
            else:
                raise
        registrar_log("user_crear", f"{nombre} ({usuario}, {'admin' if admin else 'staff'})")
        return jsonify({"ok": True})
    except Exception as e:
        if "duplicate key" in str(e) or "already exists" in str(e):
            return jsonify({"ok": False, "error": "Ese usuario ya existe"}), 409
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/users/<int:user_id>", methods=["PUT"])
@require_rol("admin")
def api_users_editar(user_id):
    data = request.get_json(silent=True) or {}
    cambios = {}

    if "usuario" in data:
        nuevo_usuario = (data.get("usuario") or "").strip().lower()
        if not nuevo_usuario:
            return jsonify({"ok": False, "error": "El usuario no puede estar vacío"}), 400
        existe = supabase.table(TABLA_USERS).select("id") \
            .eq("usuario", nuevo_usuario).execute()
        if existe.data and existe.data[0]["id"] != user_id:
            return jsonify({"ok": False, "error": "Ese usuario ya existe"}), 409
        cambios["usuario"] = nuevo_usuario

    if "nombre" in data:
        cambios["nombre"] = (data.get("nombre") or "").strip()[:60]

    if "admin" in data:
        cambios["rol"] = "admin" if data["admin"] else None

    if data.get("password"):
        cambios["password_hash"] = generate_password_hash(data["password"])

    # Solo el admin general puede mover usuarios entre colegios
    if "colegio_id" in data and es_admin_general():
        cambios["colegio_id"] = int(data["colegio_id"]) if data["colegio_id"] else None

    if not cambios:
        return jsonify({"ok": False, "error": "Nada que editar"}), 400

    prev_row = _usuario_por_id(user_id)
    if not prev_row:
        return jsonify({"ok": False, "error": "Usuario no encontrado"}), 404
    if not usuario_en_alcance(prev_row):
        return jsonify({"ok": False, "error": "Ese usuario está fuera de tu colegio"}), 403

    resp = supabase.table(TABLA_USERS).update(cambios).eq("id", user_id).execute()
    if not resp.data:
        return jsonify({"ok": False, "error": "Usuario no encontrado"}), 404

    # Actualizar session si se edita el propio usuario
    if resp.data[0]["usuario"] == session.get("usuario"):
        session["usuario"] = cambios.get("usuario", resp.data[0]["usuario"])
        if "rol" in cambios:
            session["rol"] = cambios["rol"]

    antes = prev_row
    textos = []
    if "usuario" in cambios:
        textos.append(f"usuario {antes['usuario']} → {cambios['usuario']}")
    if "nombre" in cambios:
        textos.append(f"nombre {antes.get('nombre', '') or '—'} → {cambios['nombre'] or '—'}")
    if "rol" in cambios:
        textos.append(f"admin {antes['rol'] == 'admin'} → {cambios['rol'] == 'admin'}")
    if "password_hash" in cambios:
        textos.append("contraseña actualizada")
    registrar_log("user_editar", f"id {user_id} · {', '.join(textos)}")
    return jsonify({"ok": True})


@app.route("/api/users/<int:user_id>", methods=["DELETE"])
@require_rol("admin")
def api_users_borrar(user_id):
    prev_row = _usuario_por_id(user_id)
    if not prev_row:
        return jsonify({"ok": False, "error": "Usuario no encontrado"}), 404

    if prev_row["usuario"] == session.get("usuario"):
        return jsonify({"ok": False, "error": "No podés borrar tu propio usuario"}), 400
    if not usuario_en_alcance(prev_row):
        return jsonify({"ok": False, "error": "Ese usuario está fuera de tu colegio"}), 403

    supabase.table(TABLA_USERS).delete().eq("id", user_id).execute()
    registrar_log("user_borrar", f"{prev_row.get('nombre', '')} ({prev_row['usuario']})")
    return jsonify({"ok": True})


@app.route("/api/asignar", methods=["POST"])
@require_rol("admin")
def api_asignar():
    """Asigna un usuario a un evento con un rol (vendedor/portero)."""
    data = request.get_json(silent=True) or {}
    usuario_id = data.get("usuario_id")
    evento_id = data.get("evento_id")
    rol = data.get("rol")

    if not usuario_id or not evento_id or rol not in ("vendedor", "portero"):
        return jsonify({"ok": False, "error": "usuario_id, evento_id y rol (vendedor/portero) requeridos"}), 400

    user = supabase.table(TABLA_USERS).select("usuario,colegio_id") \
        .eq("id", usuario_id).execute()
    ev = supabase.table(TABLA_EVENTOS).select("nombre,colegio_id") \
        .eq("id", evento_id).execute()
    if not user.data:
        return jsonify({"ok": False, "error": "Usuario no encontrado"}), 404
    if not ev.data:
        return jsonify({"ok": False, "error": "Evento no encontrado"}), 404

    # Admin de colegio solo asigna sobre eventos de su colegio
    if session.get("colegio_id") and \
            int(evento_id) not in {e["id"] for e in eventos_visibles()}:
        return jsonify({"ok": False, "error": "Ese evento está fuera de tu colegio"}), 403

    # Integridad: staff y evento deben ser del mismo colegio
    # (los eventos globales aceptan staff de cualquier colegio)
    u_col = user.data[0].get("colegio_id")
    e_col = ev.data[0].get("colegio_id")
    if u_col and e_col and u_col != e_col:
        return jsonify({"ok": False,
                        "error": "Ese usuario pertenece a otro colegio que el evento"}), 403

    supabase.table(TABLA_ASIGNACIONES).upsert(
        {"usuario_id": usuario_id, "evento_id": evento_id, "rol": rol},
        on_conflict="evento_id,usuario_id").execute()
    registrar_log("user_asignar", f"{user.data[0]['usuario']} → {ev.data[0]['nombre']} ({rol})")
    return jsonify({"ok": True})


@app.route("/api/asignar", methods=["DELETE"])
@require_rol("admin")
def api_desasignar():
    data = request.get_json(silent=True) or {}
    usuario_id = data.get("usuario_id")
    evento_id = data.get("evento_id")
    if not usuario_id or not evento_id:
        return jsonify({"ok": False, "error": "usuario_id y evento_id requeridos"}), 400

    # Admin de colegio solo desasigna de eventos de su colegio
    if session.get("colegio_id") and \
            int(evento_id) not in {e["id"] for e in eventos_visibles()}:
        return jsonify({"ok": False, "error": "Ese evento está fuera de tu colegio"}), 403

    supabase.table(TABLA_ASIGNACIONES).delete() \
        .eq("usuario_id", usuario_id).eq("evento_id", evento_id).execute()
    registrar_log("user_desasignar", f"usuario {usuario_id} quitado de evento {evento_id}")
    return jsonify({"ok": True})


@app.route("/api/colegios")
@require_rol("admin")
def api_colegios_listar():
    try:
        try:
            rows = supabase.table(TABLA_COLEGIOS).select(
                "id,nombre,color").order("nombre").execute()
        except Exception:
            # Columna color aún no migrada
            rows = supabase.table(TABLA_COLEGIOS).select(
                "id,nombre").order("nombre").execute()
            for c in rows.data:
                c["color"] = None
        # Conteos por colegio (tablas pequeñas, un query por tabla)
        evs = supabase.table(TABLA_EVENTOS).select("colegio_id").execute()
        uss = supabase.table(TABLA_USERS).select("colegio_id").execute()
        n_ev, n_us = {}, {}
        for r in evs.data or []:
            n_ev[r.get("colegio_id")] = n_ev.get(r.get("colegio_id"), 0) + 1
        for r in uss.data or []:
            n_us[r.get("colegio_id")] = n_us.get(r.get("colegio_id"), 0) + 1
        colegios = [{**c,
                     "eventos": n_ev.get(c["id"], 0),
                     "usuarios": n_us.get(c["id"], 0)} for c in rows.data]
        return jsonify({"ok": True, "colegios": colegios})
    except Exception:
        # Migración de colegios pendiente: lista vacía, la app sigue normal
        return jsonify({"ok": True, "colegios": []})


@app.route("/api/colegios", methods=["POST"])
@require_rol("admin")
def api_colegios_crear():
    if not es_admin_general():
        return jsonify({"ok": False,
                        "error": "Solo el admin general puede crear colegios"}), 403

    data = request.get_json(silent=True) or {}
    nombre = (data.get("nombre") or "").strip()[:80]
    if not nombre:
        return jsonify({"ok": False, "error": "Nombre del colegio requerido"}), 400

    existe = supabase.table(TABLA_COLEGIOS).select("id") \
        .eq("nombre", nombre).execute()
    if existe.data:
        return jsonify({"ok": False, "error": "Ese colegio ya existe"}), 409

    color = (data.get("color") or "").strip().lower()
    if not re.fullmatch(r"#[0-9a-f]{6}", color):
        # Color automático: siguiente de la paleta según cantidad de colegios
        existentes = supabase.table(TABLA_COLEGIOS).select("id").execute()
        PALETA = ["#d9ff3d", "#4dd4ac", "#5aa9ff", "#ff7ab6",
                  "#ffb84d", "#b78bff", "#ff6b6b", "#4dd4ff"]
        color = PALETA[(existentes.count or len(existentes.data or [])) % len(PALETA)]

    try:
        resp = supabase.table(TABLA_COLEGIOS).insert(
            {"nombre": nombre, "color": color}).execute()
        registrar_log("colegio_crear", nombre)
        return jsonify({"ok": True, "id": resp.data[0]["id"], "nombre": nombre,
                        "color": color})
    except Exception as e:
        if "color" in str(e) and ("column" in str(e) or "schema cache" in str(e)):
            # Migración de color pendiente: crear sin color
            try:
                resp = supabase.table(TABLA_COLEGIOS).insert(
                    {"nombre": nombre}).execute()
                registrar_log("colegio_crear", nombre)
                return jsonify({"ok": True, "id": resp.data[0]["id"],
                                "nombre": nombre, "color": None})
            except Exception as e2:
                return jsonify({"ok": False, "error": str(e2)}), 500
        if "does not exist" in str(e) or "schema cache" in str(e):
            return jsonify({"ok": False,
                            "error": "Ejecutá migrar_colegios.sql en Supabase primero"}), 400
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/colegios/<int:colegio_id>", methods=["PUT"])
@require_rol("admin")
def api_colegios_color(colegio_id):
    """Cambiar el color del colegio (para las gráficas)."""
    if not es_admin_general():
        return jsonify({"ok": False,
                        "error": "Solo el admin general puede editar colegios"}), 403

    data = request.get_json(silent=True) or {}
    color = (data.get("color") or "").strip().lower()
    if not re.fullmatch(r"#[0-9a-f]{6}", color):
        return jsonify({"ok": False, "error": "Color inválido (#rrggbb)"}), 400

    try:
        supabase.table(TABLA_COLEGIOS).update(
            {"color": color}).eq("id", colegio_id).execute()
        registrar_log("colegio_editar", f"colegio {colegio_id} → color {color}")
        return jsonify({"ok": True, "color": color})
    except Exception as e:
        if "color" in str(e) and ("column" in str(e) or "schema cache" in str(e)):
            return jsonify({"ok": False,
                            "error": "Ejecutá 20260823000000_color_y_integridad.sql en Supabase primero"}), 400
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/colegios/<int:colegio_id>", methods=["DELETE"])
@require_rol("admin")
def api_colegios_borrar(colegio_id):
    if not es_admin_general():
        return jsonify({"ok": False,
                        "error": "Solo el admin general puede borrar colegios"}), 403

    try:
        evs = supabase.table(TABLA_EVENTOS).select("id", count="exact") \
            .eq("colegio_id", colegio_id).execute()
        uss = supabase.table(TABLA_USERS).select("id", count="exact") \
            .eq("colegio_id", colegio_id).execute()
        en_uso = (evs.count or 0) + (uss.count or 0)
        if en_uso:
            return jsonify({"ok": False,
                            "error": f"No se puede borrar: {en_uso} elemento(s) asignados a este colegio. Reasignalos primero."}), 409

        supabase.table(TABLA_COLEGIOS).delete().eq("id", colegio_id).execute()
        registrar_log("colegio_borrar", f"colegio {colegio_id}")
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# -----------------------------------------------------------
# Entry point
# -----------------------------------------------------------
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)